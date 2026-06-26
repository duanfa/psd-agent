from __future__ import annotations

import shutil
from pathlib import Path

from fastapi import UploadFile

from .models import UploadedAsset
from .pipeline import classify_asset


def safe_filename(name: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in "._-()[] " else "_" for ch in name)
    return cleaned.strip() or "asset"


def extract_spreadsheet_text(path: Path) -> str | None:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        try:
            return path.read_text(encoding="utf-8")[:16000]
        except UnicodeDecodeError:
            return path.read_text(encoding="gb18030", errors="ignore")[:16000]
        except Exception:
            return None
    if suffix not in {".xlsx", ".xlsm"}:
        return None
    try:
        from openpyxl import load_workbook
    except Exception:
        return None

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets[:6]:
            lines.append(f"[Sheet] {sheet.title}")
            count = 0
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if values:
                    lines.append(" | ".join(values))
                    count += 1
                if count >= 120:
                    break
        return "\n".join(lines)[:16000]
    except Exception:
        return None


async def save_uploaded_assets(
    files: list[UploadFile],
    input_dir: Path,
    bucket_override: str | None = None,
) -> list[UploadedAsset]:
    input_dir.mkdir(parents=True, exist_ok=True)
    assets: list[UploadedAsset] = []
    for file in files:
        filename = safe_filename(file.filename or "asset")
        target = input_dir / filename
        with target.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        assets.append(
            UploadedAsset(
                name=filename,
                content_type=file.content_type,
                size=target.stat().st_size,
                saved_path=str(target),
                extracted_text=extract_spreadsheet_text(target),
                bucket=bucket_override or classify_asset(filename, file.content_type),
            )
        )
    return assets
