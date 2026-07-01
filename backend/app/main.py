from __future__ import annotations

import json
import mimetypes
import shutil
import uuid
from pathlib import Path

from fastapi import FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field, ValidationError

from . import database
from .defaults import (
    describe_system_model_config_source,
    load_system_model_config,
    load_workflow_defaults,
)
from .input_layers import (
    build_asset_input_layers,
    build_input_layers,
    input_layers_log_payload,
)
from .llm import LLMClient, LLMUnavailable, resolve_api_key, resolve_base_url
from .models import UploadedAsset, WorkflowArtifacts, WorkflowRequest, WorkflowResult
from .pipeline import WorkflowCancelled, classify_asset, run_pipeline
from .render import build_design_spec, write_artifacts
from .runtime import append_log, get_run_snapshot, reset_run, sanitize_for_log, set_run_state
from .wireframe import parse_wireframe_spec, wireframe_summary_text

APP_ROOT = Path(__file__).resolve().parents[1]
RUNS_ROOT = APP_ROOT / "runs"
CANCELLED_RUNS: set[str] = set()


class TrainBrandRuleRequest(BaseModel):
    brand_id: int
    asset_ids: list[int] = []
    prompt: str
    website_urls: list[str] = []
    base_version_id: int | None = None
    client_run_id: str | None = None
    training_target: str = "brand_core"


class UpdateBrandRuleMarkdownRequest(BaseModel):
    markdown: str


class DesignFeedbackRequest(BaseModel):
    feedback_type: str = "designer_edit"
    author: str = "designer"
    changes: list[dict[str, object]] = Field(default_factory=list)
    notes: str = ""


class BrandRequest(BaseModel):
    name: str
    status: str = "active"


class BrandAssetTrainingMetaRequest(BaseModel):
    training_role: str = "reference"
    include_in_training: bool = False
    quality_level: str = "normal"


class ModelTestMessage(BaseModel):
    role: str
    content: str


class ModelTestRequest(BaseModel):
    messages: list[ModelTestMessage] = Field(default_factory=list)


class ModelTestConfigResponse(BaseModel):
    provider: str
    model: str
    vision_model: str
    base_url: str | None = None
    temperature: float
    max_tokens: int
    enable_vision: bool
    max_vision_images: int
    has_api_key: bool = False
    source_path: str


class ModelTestResponse(BaseModel):
    reply: str
    provider: str
    model: str
    base_url: str | None = None

app = FastAPI(title="BrandOS AI Design Platform", version="0.3.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup() -> None:
    database.init_db()
    database.ensure_seed_data()


@app.get("/api/health")
def health() -> dict[str, object]:
    return {"status": "ok", "database": database.database_health()}


@app.post("/api/workflows/{run_id}/cancel")
def cancel_workflow(run_id: str) -> dict[str, str]:
    CANCELLED_RUNS.add(run_id)
    append_log(run_id, "Workflow", "收到用户取消请求")
    set_run_state(run_id, "cancelling", None)
    return {"status": "cancelling", "run_id": run_id}


@app.get("/api/workflows/{run_id}/logs")
def workflow_logs(run_id: str) -> dict[str, object]:
    return get_run_snapshot(run_id)


@app.get("/api/workflows/{run_id}")
def workflow_detail(run_id: str) -> dict[str, object]:
    detail = database.get_workflow_detail(run_id)
    if detail is None:
        raise HTTPException(status_code=404, detail="workflow run not found")
    return detail


@app.get("/api/workflows/{run_id}/feedback")
def workflow_feedback(run_id: str) -> dict[str, object]:
    return database.list_design_feedback(run_id)


@app.post("/api/workflows/{run_id}/feedback")
def create_workflow_feedback(
    run_id: str,
    payload: DesignFeedbackRequest,
) -> dict[str, object]:
    try:
        result = database.record_design_feedback(
            run_id=run_id,
            feedback_type=payload.feedback_type,
            author=payload.author,
            changes=payload.changes,
            notes=payload.notes,
        )
        append_log(
            run_id,
            "Feedback",
            "设计师反馈已记录，不自动覆盖品牌规则",
            result,
        )
        return result
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.get("/api/config/defaults")
def config_defaults() -> dict[str, object]:
    defaults = database.get_workflow_defaults_data()
    stages = database.get_workflow_stages_data()
    if defaults is None or stages is None:
        database.ensure_seed_data()
        defaults = database.get_workflow_defaults_data() or load_workflow_defaults()
        stages = database.get_workflow_stages_data() or database.DEFAULT_WORKFLOW_STAGES
    defaults = _merge_payload(defaults)
    return {
        "payload": defaults,
        "prompts": defaults["prompts"],
        "workflowModes": ["smart_recommend", "strict_brand"],
        "outputTypes": ["detail_page", "figma_page", "psd_file", "main_image", "banner"],
        "stages": stages,
    }


@app.post("/api/model-test", response_model=ModelTestResponse)
def model_test(payload: ModelTestRequest) -> ModelTestResponse:
    try:
        request = WorkflowRequest.model_validate(
            _merge_payload({"model_config": load_system_model_config()})
        )
    except (ValidationError, ValueError) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    messages = [
        {"role": item.role.strip(), "content": item.content.strip()}
        for item in payload.messages
        if item.content.strip()
    ]
    if not messages:
        raise HTTPException(status_code=422, detail="messages 不能为空")
    if not any(item["role"] == "user" for item in messages):
        raise HTTPException(status_code=422, detail="至少提供一条 user 消息")

    settings = request.model_settings
    settings.enable_deepagents = True
    run_id = f"model-test-{uuid.uuid4().hex[:12]}"
    reset_run(run_id)
    try:
        database.persist_run_started(run_id, request, [])
    except Exception as exc:
        print(f"[DB] persist_run_started failed: {exc}", flush=True)
    set_run_state(run_id, "running", "model_test", "模型可用性测试", "sparkles")
    llm = LLMClient(settings, run_id=run_id)

    try:
        reply = llm.invoke_messages(messages)
    except LLMUnavailable as exc:
        set_run_state(run_id, "failed", None)
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    set_run_state(run_id, "completed", None)

    return ModelTestResponse(
        reply=reply,
        provider=settings.provider,
        model=settings.model,
        base_url=resolve_base_url(settings, settings.model),
    )


@app.get("/api/model-test/config", response_model=ModelTestConfigResponse)
def model_test_config() -> ModelTestConfigResponse:
    try:
        request = WorkflowRequest.model_validate(
            _merge_payload({"model_config": load_system_model_config()})
        )
    except (ValidationError, ValueError) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    settings = request.model_settings
    return ModelTestConfigResponse(
        provider=settings.provider,
        model=settings.model,
        vision_model=settings.vision_model,
        base_url=resolve_base_url(settings, settings.model),
        temperature=settings.temperature,
        max_tokens=settings.max_tokens,
        enable_vision=settings.enable_vision,
        max_vision_images=settings.max_vision_images,
        has_api_key=bool(resolve_api_key(settings)),
        source_path=describe_system_model_config_source(),
    )


@app.get("/api/pages/dashboard")
def dashboard_page() -> dict[str, object]:
    data = database.get_dashboard_data()
    if data is None:
        database.ensure_seed_data()
        data = database.get_dashboard_data()
    if data is None:
        raise HTTPException(status_code=404, detail="dashboard data not found")
    return data


@app.get("/api/pages/brand-assets")
def brand_assets_page(
    brand_id: int | None = Query(default=None),
    folder: str | None = Query(default=None),
    status: str | None = Query(default=None),
    search: str | None = Query(default=None),
) -> dict[str, object]:
    data = database.get_brand_assets_page_data(
        brand_id=brand_id,
        folder=folder,
        status=status,
        search=search,
    )
    if data is None:
        database.ensure_seed_data()
        data = database.get_brand_assets_page_data(
            brand_id=brand_id,
            folder=folder,
            status=status,
            search=search,
        )
    if data is None:
        raise HTTPException(status_code=404, detail="brand assets data not found")
    return data


@app.post("/api/brands")
def create_brand(payload: BrandRequest) -> dict[str, object]:
    try:
        return database.create_brand(name=payload.name, status=payload.status)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.put("/api/brands/{brand_id}")
def update_brand(brand_id: int, payload: BrandRequest) -> dict[str, object]:
    try:
        return database.update_brand(brand_id=brand_id, name=payload.name, status=payload.status)
    except ValueError as exc:
        status_code = 404 if "not found" in str(exc) else 422
        raise HTTPException(status_code=status_code, detail=str(exc)) from exc


@app.delete("/api/brands/{brand_id}")
def delete_brand(brand_id: int) -> dict[str, object]:
    try:
        result = database.delete_brand(brand_id)
    except ValueError as exc:
        status_code = 404 if "not found" in str(exc) else 422
        raise HTTPException(status_code=status_code, detail=str(exc)) from exc
    for path in result.get("filePaths", []):
        try:
            file_path = Path(str(path))
            if file_path.is_file():
                file_path.unlink()
        except Exception:
            pass
    return {"id": result["id"], "deletedAssets": result["deletedAssets"]}


@app.get("/api/pages/brand-rules")
def brand_rules_page(
    brand_id: int | None = Query(default=None),
    version_id: int | None = Query(default=None),
) -> dict[str, object]:
    data = database.get_brand_rules_page_data(brand_id=brand_id, version_id=version_id)
    if data is None:
        database.ensure_seed_data()
        data = database.get_brand_rules_page_data(brand_id=brand_id, version_id=version_id)
    if data is None:
        raise HTTPException(status_code=404, detail="brand rules data not found")
    return data


@app.get("/api/brand-rules/options")
def brand_rule_options() -> dict[str, object]:
    data = database.get_brand_rule_options_data()
    if data is None:
        database.ensure_seed_data()
        data = database.get_brand_rule_options_data()
    if data is None:
        raise HTTPException(status_code=404, detail="brand rule options not found")
    return data


@app.get("/api/brand-rules/train/{run_id}/logs")
def brand_rule_train_logs(run_id: str) -> dict[str, object]:
    return get_run_snapshot(run_id)


def _brand_rule_model_result(
    run_id: str,
    payload: TrainBrandRuleRequest,
) -> dict[str, object]:
    set_run_state(run_id, "running", "prepare", "准备训练输入", "layers")
    context = database.get_brand_rule_training_context(
        brand_id=payload.brand_id,
        asset_ids=payload.asset_ids,
        base_version_id=payload.base_version_id,
        training_target=payload.training_target,
    )
    target_meta = context["targetMeta"]
    defaults = load_workflow_defaults()
    request_defaults = WorkflowRequest.model_validate(defaults)
    settings = request_defaults.model_settings
    settings.enable_deepagents = True
    settings.enable_vision = True
    llm = LLMClient(settings, run_id=run_id)
    append_log(
        run_id,
        "BrandRuleTrain",
        "复用 PSD 生成模型配置",
        {
            "provider": settings.provider,
            "model": settings.model,
            "vision_model": settings.vision_model,
            "base_url": settings.base_url,
            "enable_vision": settings.enable_vision,
            "max_vision_images": settings.max_vision_images,
            "has_api_key": bool(settings.api_key),
        },
    )

    image_paths = [
        str(asset.get("savedPath"))
        for asset in context["assets"]
        if asset.get("savedPath")
        and Path(str(asset.get("savedPath"))).is_file()
        and (
            str(asset.get("contentType") or "").startswith("image/")
            or Path(str(asset.get("savedPath"))).suffix.lower() in {".png", ".jpg", ".jpeg", ".webp", ".gif"}
        )
    ][: settings.max_vision_images]

    output_schema: dict[str, object] = {
        "markdown": f"完整 {target_meta['label']} Markdown",
        "design_rules": [{"title": "规则标题", "description": "规则说明"}],
        "layout_rules": [{"title": "布局规则标题", "description": "布局规则说明"}],
        "components": [{"title": "组件标题", "description": "组件说明"}],
        "source_assets": [{"title": "素材名称", "description": "从该素材提取出的规范"}],
    }
    if payload.training_target == "detail_page_layout":
        output_schema.update(
            {
                "layout_schema": {
                    "schema_version": "brandos_layout_schema.v1",
                    "canvas": {"width": 790, "height_mode": "auto"},
                    "sections": [],
                    "image_slots": [],
                    "text_layers": [],
                    "component_templates": [],
                    "global_constraints": {},
                },
                "image_slots": [],
            }
        )
    else:
        output_schema["visual_tokens"] = {
            "colors": [],
            "typography": {},
            "spacing": {},
            "tone": "",
        }

    model_payload = {
        "brand": context["brand"],
        "training_target": payload.training_target,
        "training_target_label": target_meta["label"],
        "selected_assets": [
            {
                "id": asset["id"],
                "name": asset["name"],
                "folder": asset["folder"],
                "contentType": asset["contentType"],
                "source": asset["source"],
                "extractedText": str(asset.get("extractedText") or "")[:2000],
                "metadata": asset.get("metadata") or {},
            }
            for asset in context["assets"]
        ],
        "base_rule": context["baseRule"],
        "linked_core_rule": context["linkedCoreRule"],
        "website_urls": payload.website_urls,
        "training_prompt": payload.prompt,
        "output_schema": output_schema,
    }
    append_log(run_id, "BrandRuleTrain", "训练输入上下文", model_payload)

    set_run_state(run_id, "running", "vision_model", f"多模态模型提取{target_meta['label']}", "sparkles")
    if payload.training_target == "detail_page_layout":
        system_prompt = (
            "你是 BrandOS 的详情页布局规则训练 Agent。必须基于用户勾选的详情页素材、素材文本、"
            "可见图片内容、可选历史详情页规则和关联品牌核心规则，提取可复用的商品详情页 Derived Rule。"
            "如果提供历史版本，需要明确区分保留规则和新增补充。请输出可直接入库的 JSON。"
        )
        user_prompt = (
            "请根据以下 JSON 上下文生成详情页布局规则。要求：\n"
            "1. 认真阅读 selected_assets，每个素材都要在 source_assets 中有提炼结果。\n"
            "2. 若 base_rule 不为空，叠加历史详情页规则并说明新增模块；若为空，创建全新详情页规则。\n"
            "3. markdown 要包含：训练输入、素材摘要、页面结构、文字排版层级、图片放置区域、留白与栅格、组件模板、禁用项。\n"
            "4. layout_rules 必须明确描述主图区、卖点图区、参数/对比区、CTA 区等位置或比例关系。\n"
            "5. 必须额外输出可执行 layout_schema，包含 sections、image_slots、text_layers、component_templates；"
            "每个 section 尽量给出 id、name、role、component_type、order、x/y/w/h、background、required_image_slots、required_text_fields；"
            "每个 slot 尽量给出 id、section_id、role、asset_type、x/y/w/h、fit/crop、priority、required。\n"
            "6. 如果 selected_assets.metadata.wireframe_spec 存在，优先基于其中的 drawing object、cell box、merged range、media path 抽取 layout_schema。\n"
            "7. design_rules/layout_rules/components/source_assets 都必须是数组，每项包含 title 和 description；"
            "layout_schema 和 image_slots 作为 JSON 字段单独输出。\n\n"
            f"{json.dumps(model_payload, ensure_ascii=False, indent=2)}"
        )
    else:
        system_prompt = (
            "你是 BrandOS 的品牌设计规范训练 Agent。必须基于用户勾选的素材、素材文本、官网 URL、"
            "可见图片内容和可选历史品牌规则，提取可复用的品牌级 Core Rule。"
            "如果提供历史版本，需要明确区分保留规则和新增补充。请输出可直接入库的 JSON。"
        )
        user_prompt = (
            "请根据以下 JSON 上下文生成品牌设计规范。要求：\n"
            "1. 认真阅读 selected_assets，每个素材都要在 source_assets 中有提炼结果。\n"
            "2. 若 base_rule 不为空，叠加历史品牌规则并说明新增规范；若为空，创建全新品牌核心规则。\n"
            "3. markdown 要包含：训练输入、素材摘要、品牌视觉规范、色彩/字体/语气、禁用项、供页面规则继承的品牌级布局倾向。\n"
            "4. 不要把具体商品详情页模块当成唯一模板写死在 Core Rule 中。\n"
            "5. 必须额外输出 visual_tokens，包含 colors、typography、spacing、tone、do/dont 等可执行视觉 Token。\n"
            "6. design_rules/layout_rules/components/source_assets 都必须是数组，每项包含 title 和 description。\n\n"
            f"{json.dumps(model_payload, ensure_ascii=False, indent=2)}"
        )
    try:
        result = llm.invoke_vision_json(system_prompt, user_prompt, image_paths)
    except LLMUnavailable as exc:
        append_log(run_id, "BrandRuleTrain", "多模态模型训练失败", str(exc))
        set_run_state(run_id, "failed", "vision_model", "多模态模型提取品牌规范", "sparkles")
        raise
    append_log(run_id, "BrandRuleTrain", "模型解析后的品牌规则 JSON", result)
    return result


@app.post("/api/brand-rules/train")
def train_brand_rules(payload: TrainBrandRuleRequest) -> dict[str, object]:
    run_id = payload.client_run_id or f"brand-rule-{uuid.uuid4()}"
    reset_run(run_id)
    try:
        model_result = _brand_rule_model_result(run_id, payload)
        set_run_state(run_id, "running", "persist", "保存训练版本", "check-circle")
        result = database.train_brand_rule_version(
            brand_id=payload.brand_id,
            asset_ids=payload.asset_ids,
            prompt=payload.prompt,
            website_urls=payload.website_urls,
            base_version_id=payload.base_version_id,
            model_result=model_result,
            training_target=payload.training_target,
        )
        append_log(run_id, "BrandRuleTrain", "品牌规则版本已保存", result)
        return result
    except LLMUnavailable as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    except ValueError as exc:
        append_log(run_id, "BrandRuleTrain", "训练失败", str(exc))
        set_run_state(run_id, "failed", None)
        status_code = 404 if "not found" in str(exc) else 422
        raise HTTPException(status_code=status_code, detail=str(exc)) from exc
    finally:
        snapshot = get_run_snapshot(run_id)
        if snapshot.get("status") == "running":
            set_run_state(run_id, "completed", None)


@app.put("/api/brand-rules/{rule_id}/markdown")
def update_brand_rule_markdown(
    rule_id: int,
    payload: UpdateBrandRuleMarkdownRequest,
) -> dict[str, object]:
    try:
        return database.update_brand_rule_markdown(rule_id, payload.markdown)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.post("/api/brand-rules/{rule_id}/publish")
def publish_brand_rule(rule_id: int) -> dict[str, object]:
    try:
        return database.publish_brand_rule_version(rule_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.post("/api/brand-rules/{rule_id}/rollback")
def rollback_brand_rule(rule_id: int) -> dict[str, object]:
    try:
        return database.rollback_brand_rule_version(rule_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.get("/api/brand-rules/diff")
def diff_brand_rules(
    base_rule_id: int = Query(...),
    compare_rule_id: int = Query(...),
) -> dict[str, object]:
    try:
        return database.diff_brand_rule_versions(base_rule_id, compare_rule_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.delete("/api/brand-rules/{rule_id}")
def delete_brand_rule_version(rule_id: int) -> dict[str, object]:
    try:
        return database.delete_brand_rule_version(rule_id=rule_id)
    except ValueError as exc:
        status_code = 404 if "not found" in str(exc) else 422
        raise HTTPException(status_code=status_code, detail=str(exc)) from exc


@app.get("/api/pages/products")
def products_page() -> dict[str, object]:
    data = database.get_products_data()
    if data is None:
        database.ensure_seed_data()
        data = database.get_products_data()
    if data is None:
        raise HTTPException(status_code=404, detail="products data not found")
    return data


@app.get("/api/pages/design-tasks")
def design_tasks_page(
    brand: str | None = Query(default=None),
    status: str | None = Query(default=None),
    task_type: str | None = Query(default=None),
    search: str | None = Query(default=None),
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
) -> dict[str, object]:
    data = database.get_design_tasks_page_data(
        brand=brand,
        status=status,
        task_type=task_type,
        search=search,
        limit=limit,
        offset=offset,
    )
    if data is None:
        database.ensure_seed_data()
        data = database.get_design_tasks_page_data(
            brand=brand,
            status=status,
            task_type=task_type,
            search=search,
            limit=limit,
            offset=offset,
        )
    if data is None:
        raise HTTPException(status_code=404, detail="design tasks data not found")
    return data


def _safe_filename(name: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in "._-()[] " else "_" for ch in name)
    return cleaned.strip() or "asset"


def _safe_run_id(run_id: str | None) -> str:
    if not run_id:
        return uuid.uuid4().hex
    cleaned = "".join(ch if ch.isalnum() or ch in "-_" else "" for ch in run_id)
    return cleaned[:80] or uuid.uuid4().hex


def _extract_spreadsheet_text(path: Path) -> str | None:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        try:
            return path.read_text(encoding="utf-8")[:16000]
        except UnicodeDecodeError:
            return path.read_text(encoding="gb18030", errors="ignore")[:16000]
        except Exception:
            return None
    if suffix == ".xls":
        try:
            import xlrd  # type: ignore
        except Exception:
            return None

        try:
            workbook = xlrd.open_workbook(str(path))
            lines: list[str] = []
            for sheet in workbook.sheets()[:6]:
                lines.append(f"[Sheet] {sheet.name}")
                count = 0
                for row_idx in range(sheet.nrows):
                    values = [
                        str(value).strip()
                        for value in sheet.row_values(row_idx)
                        if value not in (None, "")
                    ]
                    if values:
                        lines.append(" | ".join(values))
                        count += 1
                    if count >= 120:
                        break
            return "\n".join(lines)[:16000]
        except Exception:
            return None
    if suffix not in {".xlsx", ".xlsm"}:
        return None
    try:
        from openpyxl import load_workbook
    except Exception:
        return None

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets[:6]:
            lines.append(f"[Sheet] {sheet.title}")
            count = 0
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if values:
                    lines.append(" | ".join(values))
                    count += 1
                if count >= 120:
                    break
        wireframe = wireframe_summary_text(parse_wireframe_spec(path))
        return "\n".join([*lines, wireframe]).strip()[:24000]
    except Exception:
        return None


def _save_assets(
    files: list[UploadFile],
    input_dir: Path,
    bucket_override: str | None = None,
) -> list[UploadedAsset]:
    input_dir.mkdir(parents=True, exist_ok=True)
    assets: list[UploadedAsset] = []
    for file in files:
        filename = _safe_filename(file.filename or "asset")
        target = input_dir / filename
        with target.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        wireframe_spec = parse_wireframe_spec(target)
        extracted_text = _extract_spreadsheet_text(target)
        metadata: dict[str, object] = {}
        if wireframe_spec:
            metadata["wireframe_spec"] = wireframe_spec
            metadata["wireframe_summary"] = {
                "sheet_count": len(wireframe_spec.get("sheets") or []),
                "media_count": wireframe_spec.get("media_count", 0),
                "drawing_object_count": wireframe_spec.get("drawing_object_count", 0),
            }
        input_layers = build_asset_input_layers(filename, extracted_text, metadata)
        if input_layers:
            metadata["input_layers"] = input_layers
        bucket = bucket_override or classify_asset(filename, file.content_type)
        assets.append(
            UploadedAsset(
                name=filename,
                content_type=file.content_type,
                size=target.stat().st_size,
                saved_path=str(target),
                extracted_text=extracted_text,
                metadata=metadata,
                bucket=bucket,
            )
        )
    return assets


def _merge_payload(incoming: dict) -> dict:
    data = load_workflow_defaults()
    for key, value in incoming.items():
        if key == "model_config":
            continue
        if isinstance(data.get(key), dict) and isinstance(value, dict):
            merged = dict(data[key])
            merged.update(value)
            data[key] = merged
        else:
            data[key] = value
    data["model_config"] = load_system_model_config()
    return data


def _looks_like_example_brief(brief: str, product_name: str) -> bool:
    text = (brief or "").strip()
    if not text:
        return False
    product = (product_name or "").strip()
    example_markers = ("商品类型：香薰机", "商品类型：电脑包", "轻量通勤", "静音运行", "持久扩香")
    has_example_marker = any(marker in text for marker in example_markers)
    return has_example_marker and (not product or product not in text)


def _validate_workflow_rule_selection(
    request: WorkflowRequest,
    selected_rule_context: dict[str, object],
) -> None:
    core_rule = dict(selected_rule_context.get("coreRule") or {})
    detail_rule = dict(selected_rule_context.get("detailPageRule") or {})
    produces_detail_page = any(item.value == "detail_page" for item in request.output_types)

    if request.selected_core_rule_id and not core_rule:
        raise HTTPException(status_code=422, detail="所选品牌 Core Rule 未命中，请重新选择后再生成")
    if request.selected_detail_page_rule_id and not detail_rule:
        raise HTTPException(status_code=422, detail="所选详情页 Derived Rule 未命中，请重新选择后再生成")
    if produces_detail_page and core_rule and not detail_rule:
        raise HTTPException(
            status_code=422,
            detail=(
                "当前品牌已命中 Core Rule，但没有已发布的详情页布局规范。"
                "请先在规则训练中选择“详情页布局规范”目标，训练并发布 active Derived Rule 后再生成详情页。"
            ),
        )

    if request.workflow_mode.value != "strict_brand":
        return

    missing: list[str] = []
    if not core_rule:
        missing.append("品牌 Core Rule")
    if not detail_rule:
        missing.append("详情页 Derived Rule")
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"strict_brand 模式要求命中 {' 和 '.join(missing)}，当前无法继续生成",
        )

    inactive: list[str] = []
    if str(core_rule.get("status") or "") != "active":
        inactive.append("品牌 Core Rule 不是已发布生效版本")
    if str(detail_rule.get("status") or "") != "active":
        inactive.append("详情页 Derived Rule 不是已发布生效版本")
    if inactive:
        raise HTTPException(status_code=422, detail="；".join(inactive))

    detail_layout_rules = detail_rule.get("layoutRules") or detail_rule.get("layout_rules") or []
    detail_components = detail_rule.get("components") or []
    if not detail_layout_rules and not detail_components:
        raise HTTPException(
            status_code=422,
            detail="strict_brand 模式要求命中的详情页 Derived Rule 至少包含布局规则或组件规则",
        )


@app.post("/api/brand-assets/upload")
async def upload_brand_assets(
    brand_id: int = Form(...),
    name: str = Form(default=""),
    folder: str = Form(...),
    source: str = Form(default=""),
    files: list[UploadFile] = File(default=[]),
) -> dict[str, object]:
    if not files:
        raise HTTPException(status_code=422, detail="请至少上传一个文件")
    media_dir = APP_ROOT / "media" / "brand-assets" / str(brand_id)
    saved_assets = _save_assets(files, media_dir, bucket_override="brand_asset")
    try:
        created = database.create_brand_assets(
            brand_id=brand_id,
            name=name,
            folder=folder,
            source=source,
            files=[asset.model_dump() for asset in saved_assets],
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return {"created": created, "count": len(created)}


@app.delete("/api/brand-assets/{asset_id}")
def delete_brand_asset(asset_id: int) -> dict[str, object]:
    try:
        result = database.delete_brand_asset(asset_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    saved_path = Path(str(result.get("savedPath") or ""))
    try:
        if saved_path.is_file():
            saved_path.unlink()
    except Exception:
        pass
    return {"id": result["id"], "brandId": result["brandId"]}


@app.put("/api/brand-assets/{asset_id}")
def update_brand_asset_training_meta(
    asset_id: int,
    payload: BrandAssetTrainingMetaRequest,
) -> dict[str, object]:
    try:
        return database.update_brand_asset_training_meta(
            asset_id=asset_id,
            training_role=payload.training_role,
            include_in_training=payload.include_in_training,
            quality_level=payload.quality_level,
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


def _asset_preview_type(content_type: str, path: Path | None) -> str:
    lower_name = path.name.lower() if path else ""
    if content_type.startswith("image/") or lower_name.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif")):
        return "image"
    if content_type == "application/pdf" or lower_name.endswith(".pdf"):
        return "pdf"
    if content_type.startswith("text/") or lower_name.endswith((".txt", ".md", ".csv", ".html", ".json")):
        return "text"
    return "unknown"


@app.get("/api/brand-assets/{asset_id}/preview")
def preview_brand_asset(asset_id: int) -> dict[str, object]:
    asset = database.get_brand_asset(asset_id)
    if asset is None:
        raise HTTPException(status_code=404, detail="asset not found")
    path = Path(str(asset.get("savedPath") or ""))
    file_exists = bool(asset.get("savedPath")) and path.is_file()
    content_type = str(asset.get("contentType") or "")
    if file_exists and not content_type:
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    preview_type = _asset_preview_type(content_type, path if file_exists else None)
    text_preview = str(asset.get("extractedText") or "")
    if file_exists and preview_type == "text" and not text_preview:
        try:
            text_preview = path.read_text(encoding="utf-8", errors="ignore")[:12000]
        except Exception:
            text_preview = ""
    return {
        **asset,
        "contentType": content_type,
        "fileExists": file_exists,
        "previewType": preview_type if file_exists or text_preview else "metadata",
        "fileUrl": f"/api/brand-assets/{asset_id}/file" if file_exists else "",
        "textPreview": text_preview[:12000],
    }


@app.get("/api/brand-assets/{asset_id}/file")
def download_brand_asset_file(asset_id: int) -> FileResponse:
    asset = database.get_brand_asset(asset_id)
    if asset is None:
        raise HTTPException(status_code=404, detail="asset not found")
    path = Path(str(asset.get("savedPath") or ""))
    if not path.is_file():
        raise HTTPException(status_code=404, detail="asset file not found")
    media_type = str(asset.get("contentType") or "") or mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    return FileResponse(path, media_type=media_type, filename=path.name)


@app.post("/api/workflows/generate", response_model=WorkflowResult)
def generate_workflow(
    payload: str = Form(...),
    client_run_id: str | None = Form(default=None),
    files: list[UploadFile] = File(default=[]),
    brief_files: list[UploadFile] = File(default=[]),
    reference_images: list[UploadFile] = File(default=[]),
) -> WorkflowResult:
    try:
        incoming = json.loads(payload)
        if not isinstance(incoming, dict):
            raise ValueError("payload 必须是 JSON 对象")
        request = WorkflowRequest.model_validate(_merge_payload(incoming))
    except (json.JSONDecodeError, ValidationError, ValueError) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    try:
        selected_rule_context = database.get_workflow_rule_context(
            core_rule_id=request.selected_core_rule_id,
            detail_page_rule_id=request.selected_detail_page_rule_id,
            brand_name=request.brand_name,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    if selected_rule_context.get("brandName"):
        request.brand_name = str(selected_rule_context["brandName"])
    _validate_workflow_rule_selection(request, selected_rule_context)

    feedback_constraints = database.get_feedback_constraints_for_request(
        brand_name=request.brand_name,
        product_name=request.product_name,
        feedback_scope=request.requirement_constraints.feedback_scope,
        feedback_run_id=request.requirement_constraints.feedback_run_id,
    )

    run_id = _safe_run_id(client_run_id)
    CANCELLED_RUNS.discard(run_id)
    run_dir = RUNS_ROOT / run_id
    input_dir = run_dir / "inputs"
    output_dir = run_dir / "outputs"
    assets = [
        *_save_assets(files, input_dir / "assets"),
        *_save_assets(brief_files, input_dir / "brief", bucket_override="brief"),
        *_save_assets(
            reference_images,
            input_dir / "reference_images",
            bucket_override="reference_image",
        ),
    ]

    has_layered_brief_input = any(
        asset.bucket == "brief" and ((asset.extracted_text or "").strip() or asset.metadata.get("wireframe_spec"))
        for asset in assets
    )
    if has_layered_brief_input and _looks_like_example_brief(request.product_brief, request.product_name):
        append_log(
            run_id,
            "Workflow",
            "检测到示例 Product Brief 与当前商品不匹配，已在构建分层输入前清空默认示例 brief",
            {"removed_brief": request.product_brief[:500], "product_name": request.product_name},
        )
        request.product_brief = ""

    layered_input = build_input_layers(request.product_brief, assets)
    if layered_input.get("brief_asset_count") or layered_input.get("wireframe_asset_count"):
        append_log(
            run_id,
            "Workflow",
            "Brief Excel / wireframe 已转为分层输入，不再直接拼接原文进 brief",
            input_layers_log_payload(layered_input),
        )

    try:
        stages, ctx = run_pipeline(
            request,
            assets,
            run_id=run_id,
            cancel_checker=lambda: run_id in CANCELLED_RUNS,
            selected_rule_context=selected_rule_context,
            feedback_constraints=feedback_constraints,
        )
    except WorkflowCancelled as exc:
        CANCELLED_RUNS.discard(run_id)
        set_run_state(run_id, "cancelled", None)
        raise HTTPException(status_code=499, detail=str(exc)) from exc
    except ValueError as exc:
        CANCELLED_RUNS.discard(run_id)
        set_run_state(run_id, "failed", None)
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception:
        CANCELLED_RUNS.discard(run_id)
        set_run_state(run_id, "failed", None)
        raise
    CANCELLED_RUNS.discard(run_id)
    spec = build_design_spec(ctx)
    artifact_paths = write_artifacts(output_dir, spec, ctx)

    used_model = any(stage.used_model for stage in stages)
    agent_report = "\n\n".join(ctx.report_parts)
    result_state = spec.get("result_state") if isinstance(spec.get("result_state"), dict) else {}
    delivery_status = str(result_state.get("delivery_status") or "review_only")
    if delivery_status == "ready":
        status = "completed"
    else:
        status = "fallback_completed"
    summary = (
        "BrandOS 设计任务已完成："
        f"结果等级 {result_state.get('tier') or '低保真草稿'}，"
        f"导出状态 {artifact_paths.get('export_status') or 'review_only_bundle'}。"
    )
    try:
        database.persist_run_completed(
            run_id=run_id,
            status=status,
            summary=summary,
            used_deepagents=used_model,
            agent_report=agent_report,
            design_spec=spec,
            artifact_paths=artifact_paths,
            output_dir=str(output_dir),
            warnings=ctx.warnings,
        )
    except Exception as exc:
        append_log(run_id, "Database", f"MySQL 持久化最终结果失败：{exc}")

    return WorkflowResult(
        run_id=run_id,
        status=status,
        summary=summary,
        used_deepagents=used_model,
        stages=sanitize_for_log(stages),
        agent_report=agent_report,
        design_spec=sanitize_for_log(spec),
        artifacts=WorkflowArtifacts(
            run_id=run_id,
            output_dir=str(output_dir),
            **artifact_paths,
        ),
        result_state=sanitize_for_log(spec.get("result_state") or {}),
        export_review=sanitize_for_log(spec.get("export_review") or {}),
        assets=sanitize_for_log(assets),
        warnings=ctx.warnings,
    )


@app.get("/api/workflows/{run_id}/artifacts/{name}")
def download_artifact(run_id: str, name: str) -> FileResponse:
    allowed = {
        "preview.svg": "image/svg+xml",
        "design_spec.json": "application/json",
        "output_metadata.json": "application/json",
        "create_detail_page.jsx": "text/plain",
        "create_figma_page.ts": "text/plain",
        "editable_detail_page.html": "text/html",
        "README.md": "text/markdown",
    }
    if name not in allowed:
        raise HTTPException(status_code=404, detail="artifact not found")
    path = RUNS_ROOT / run_id / "outputs" / name
    if not path.is_file():
        raise HTTPException(status_code=404, detail="artifact not found")
    return FileResponse(path, media_type=allowed[name], filename=name)
