from __future__ import annotations

import json
import posixpath
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET


EMU_PER_PIXEL = 9525
DEFAULT_COL_WIDTH = 64
DEFAULT_ROW_HEIGHT = 21

NS = {
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}


def _col_width_to_px(width: float | int | None) -> int:
    if width is None:
        return DEFAULT_COL_WIDTH
    try:
        return max(8, int(float(width) * 7 + 5))
    except (TypeError, ValueError):
        return DEFAULT_COL_WIDTH


def _row_height_to_px(height: float | int | None) -> int:
    if height is None:
        return DEFAULT_ROW_HEIGHT
    try:
        return max(8, int(float(height) * 96 / 72))
    except (TypeError, ValueError):
        return DEFAULT_ROW_HEIGHT


def _safe_text(value: Any) -> str:
    return str(value).strip().replace("\n", " / ") if value not in (None, "") else ""


def _safe_json_value(value: Any) -> str | int | float | bool | None:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _rel_targets(zip_file: zipfile.ZipFile, rels_path: str) -> dict[str, str]:
    if rels_path not in zip_file.namelist():
        return {}
    root = ET.fromstring(zip_file.read(rels_path))
    result: dict[str, str] = {}
    base_dir = posixpath.dirname(rels_path.replace("_rels/", "").replace(".rels", ""))
    for rel in root.findall("rel:Relationship", NS):
        rel_id = rel.attrib.get("Id", "")
        target = rel.attrib.get("Target", "")
        if not rel_id or not target:
            continue
        if not target.startswith("/"):
            target = posixpath.normpath(posixpath.join(base_dir, target))
        else:
            target = target.lstrip("/")
        result[rel_id] = target
    return result


def _node_int(node: ET.Element | None, name: str) -> int:
    if node is None:
        return 0
    value = node.findtext(f"xdr:{name}", "0", NS)
    try:
        return int(value or 0)
    except ValueError:
        return 0


def _position_from_marker(
    marker: ET.Element | None,
    col_offsets: list[int],
    row_offsets: list[int],
) -> dict[str, int]:
    col = _node_int(marker, "col")
    row = _node_int(marker, "row")
    col_off = _node_int(marker, "colOff")
    row_off = _node_int(marker, "rowOff")
    x = (col_offsets[col] if col < len(col_offsets) else col * DEFAULT_COL_WIDTH) + int(col_off / EMU_PER_PIXEL)
    y = (row_offsets[row] if row < len(row_offsets) else row * DEFAULT_ROW_HEIGHT) + int(row_off / EMU_PER_PIXEL)
    return {"row": row + 1, "col": col + 1, "x": x, "y": y}


def _shape_name(anchor: ET.Element) -> str:
    c_nv_pr = anchor.find(".//xdr:cNvPr", NS)
    return c_nv_pr.attrib.get("name", "") if c_nv_pr is not None else ""


def _shape_text(anchor: ET.Element) -> str:
    texts = [
        item.text.strip()
        for item in anchor.findall(".//a:t", NS)
        if item.text and item.text.strip()
    ]
    return " ".join(texts)


def _drawing_objects(
    zip_file: zipfile.ZipFile,
    drawing_path: str,
    col_offsets: list[int],
    row_offsets: list[int],
    limit: int = 160,
) -> list[dict[str, Any]]:
    if drawing_path not in zip_file.namelist():
        return []
    rels_path = posixpath.join(
        posixpath.dirname(drawing_path), "_rels", posixpath.basename(drawing_path) + ".rels"
    )
    rels = _rel_targets(zip_file, rels_path)
    root = ET.fromstring(zip_file.read(drawing_path))
    objects: list[dict[str, Any]] = []
    for anchor in list(root)[:limit]:
        start = _position_from_marker(anchor.find("xdr:from", NS), col_offsets, row_offsets)
        end_marker = anchor.find("xdr:to", NS)
        end = _position_from_marker(end_marker, col_offsets, row_offsets) if end_marker is not None else start
        blip = anchor.find(".//a:blip", NS)
        rel_id = blip.attrib.get(f"{{{NS['r']}}}embed", "") if blip is not None else ""
        media_path = rels.get(rel_id, "")
        obj_type = "image" if media_path else "shape"
        x = start["x"]
        y = start["y"]
        w = max(1, end["x"] - x)
        h = max(1, end["y"] - y)
        objects.append(
            {
                "type": obj_type,
                "name": _shape_name(anchor) or f"{obj_type}_{len(objects) + 1}",
                "text": _shape_text(anchor),
                "cell_anchor": {
                    "from": {"row": start["row"], "col": start["col"]},
                    "to": {"row": end["row"], "col": end["col"]},
                },
                "box": {"x": x, "y": y, "w": w, "h": h},
                "media_path": media_path,
                "relationship_id": rel_id,
            }
        )
    return objects


def _sheet_drawing_path(zip_file: zipfile.ZipFile, sheet_index: int) -> str:
    sheet_path = f"xl/worksheets/sheet{sheet_index}.xml"
    rels = _rel_targets(zip_file, f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels")
    if sheet_path not in zip_file.namelist():
        return ""
    try:
        root = ET.fromstring(zip_file.read(sheet_path))
    except Exception:
        return ""
    drawing = root.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}drawing")
    if drawing is None:
        return ""
    rel_id = drawing.attrib.get(f"{{{NS['r']}}}id", "")
    return rels.get(rel_id, "")


def parse_wireframe_spec(path: Path, max_cells_per_sheet: int = 260) -> dict[str, Any] | None:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return None
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return None

    try:
        workbook = load_workbook(path, read_only=False, data_only=True)
    except Exception:
        return None

    spec: dict[str, Any] = {
        "schema_version": "wireframe_spec.v1",
        "source_file": path.name,
        "sheets": [],
        "media_count": 0,
        "drawing_object_count": 0,
    }

    try:
        zip_file_cm = zipfile.ZipFile(path)
    except Exception:
        zip_file_cm = None

    with zip_file_cm if zip_file_cm is not None else _NullZip() as zip_file:
        if zip_file is not None:
            spec["media_count"] = len([name for name in zip_file.namelist() if name.startswith("xl/media/") and not name.endswith("/")])

        for sheet_index, sheet in enumerate(workbook.worksheets[:8], start=1):
            col_widths = [
                _col_width_to_px(sheet.column_dimensions[get_column_letter(i)].width)
                for i in range(1, sheet.max_column + 1)
            ]
            row_heights = [
                _row_height_to_px(sheet.row_dimensions[i].height)
                for i in range(1, sheet.max_row + 1)
            ]
            col_offsets = [0]
            for width in col_widths:
                col_offsets.append(col_offsets[-1] + width)
            row_offsets = [0]
            for height in row_heights:
                row_offsets.append(row_offsets[-1] + height)

            cells: list[dict[str, Any]] = []
            for row in sheet.iter_rows():
                for cell in row:
                    text = _safe_text(cell.value)
                    if not text:
                        continue
                    row_idx = int(cell.row)
                    col_idx = int(cell.column)
                    cells.append(
                        {
                            "address": cell.coordinate,
                            "row": row_idx,
                            "col": col_idx,
                            "text": text[:500],
                            "box": {
                                "x": col_offsets[col_idx - 1],
                                "y": row_offsets[row_idx - 1],
                                "w": col_widths[col_idx - 1],
                                "h": row_heights[row_idx - 1],
                            },
                            "style": {
                                "fill": _safe_json_value(getattr(cell.fill.fgColor, "rgb", None)),
                                "font": _safe_json_value(cell.font.name),
                                "font_size": _safe_json_value(cell.font.sz),
                                "bold": bool(cell.font.bold),
                                "align": _safe_json_value(cell.alignment.horizontal),
                            },
                        }
                    )
                    if len(cells) >= max_cells_per_sheet:
                        break
                if len(cells) >= max_cells_per_sheet:
                    break

            drawing_objects: list[dict[str, Any]] = []
            drawing_path = _sheet_drawing_path(zip_file, sheet_index) if zip_file is not None else ""
            if drawing_path and zip_file is not None:
                drawing_objects = _drawing_objects(zip_file, drawing_path, col_offsets, row_offsets)
                spec["drawing_object_count"] += len(drawing_objects)

            spec["sheets"].append(
                {
                    "name": sheet.title,
                    "dimensions": {
                        "rows": sheet.max_row,
                        "cols": sheet.max_column,
                        "width_px": col_offsets[-1],
                        "height_px": row_offsets[-1],
                    },
                    "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
                    "column_widths_px": col_widths[:80],
                    "row_heights_px": row_heights[:240],
                    "cells": cells,
                    "drawing_path": drawing_path,
                    "objects": drawing_objects,
                }
            )

    return spec


class _NullZip:
    def __enter__(self) -> None:
        return None

    def __exit__(self, *_args: Any) -> None:
        return None


def wireframe_summary_text(spec: dict[str, Any] | None, max_chars: int = 8000) -> str:
    if not spec:
        return ""
    lines = [
        "[WireframeSpec]",
        f"source_file={spec.get('source_file')}",
        f"media_count={spec.get('media_count', 0)} drawing_object_count={spec.get('drawing_object_count', 0)}",
    ]
    for sheet in spec.get("sheets", [])[:6]:
        dims = sheet.get("dimensions", {})
        lines.append(
            f"[SheetWireframe] {sheet.get('name')} {dims.get('cols')}x{dims.get('rows')} "
            f"{dims.get('width_px')}x{dims.get('height_px')}px "
            f"merged={len(sheet.get('merged_ranges') or [])} objects={len(sheet.get('objects') or [])}"
        )
        for obj in (sheet.get("objects") or [])[:28]:
            box = obj.get("box") or {}
            anchor = obj.get("cell_anchor") or {}
            media = f" media={obj.get('media_path')}" if obj.get("media_path") else ""
            text = f" text={obj.get('text')}" if obj.get("text") else ""
            lines.append(
                f"- {obj.get('type')} {obj.get('name')} box={box.get('x')},{box.get('y')},{box.get('w')},{box.get('h')} "
                f"anchor={anchor}{media}{text}"
            )
        for cell in (sheet.get("cells") or [])[:40]:
            box = cell.get("box") or {}
            lines.append(
                f"- cell {cell.get('address')} box={box.get('x')},{box.get('y')},{box.get('w')},{box.get('h')} text={cell.get('text')}"
            )
    return "\n".join(lines)[:max_chars]


def compact_wireframe_json(spec: dict[str, Any] | None, max_chars: int = 12000) -> str:
    if not spec:
        return ""
    return json.dumps(spec, ensure_ascii=False, separators=(",", ":"))[:max_chars]
